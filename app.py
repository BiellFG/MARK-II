import os
import json
from datetime import datetime
from flask import Flask, render_template, request
import openpyxl

app = Flask(__name__, template_folder='Templates', static_folder='Static')

DIRETORIO_ATUAL = os.path.dirname(os.path.abspath(__file__))
CAMINHO_EXCEL = os.path.join(DIRETORIO_ATUAL, "Controle Financeiro 2026.xlsx")

MESES_ABREV = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
MESES_NOMES = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
               'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

# Cores para os gráficos
CORES = [
    '#6366f1', '#8b5cf6', '#a855f7', '#d946ef', '#ec4899',
    '#f43f5e', '#ef4444', '#f97316', '#eab308', '#22c55e',
    '#06b6d4', '#3b82f6', '#64748b', '#14b8a6', '#84cc16'
]


def formata_real(valor):
    try:
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return f"R$ {valor}"


def formata_real_curto(valor):
    """Formata valor em formato curto (ex: R$ 8.152)."""
    try:
        if valor >= 1000:
            return f"R$ {valor:,.0f}".replace(",", ".")
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return f"R$ {valor}"


def _para_numero(valor):
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if texto == '-' or texto == '':
        return 0.0
    try:
        return float(texto.replace(',', '.'))
    except ValueError:
        return 0.0


def carregar_dados():
    if not os.path.exists(CAMINHO_EXCEL):
        raise FileNotFoundError(
            "Arquivo 'Controle Financeiro 2026.xlsx' não encontrado no servidor."
        )
    wb = openpyxl.load_workbook(CAMINHO_EXCEL, data_only=True)
    ws = wb["Dados"]
    dados = []
    for row_idx in range(11, ws.max_row + 1):
        nome_despesa = ws.cell(row=row_idx, column=2).value
        if nome_despesa is None:
            continue
        nome_despesa = str(nome_despesa).strip()
        if nome_despesa.upper() == 'TOTAL' or nome_despesa == '':
            continue
        item = {'Despesa': nome_despesa}
        for i, mes in enumerate(MESES_ABREV):
            valor_celula = ws.cell(row=row_idx, column=i + 3).value
            item[mes] = _para_numero(valor_celula)
        dados.append(item)
    wb.close()
    return dados


def obter_mes_atual():
    idx = datetime.now().month - 1
    return MESES_ABREV[idx]


def agrupar_categorias(dados):
    """
    Agrupa despesas por categorias. Primeiro tenta ler da planilha 'Categorias',
    se não existir, usa mapeamento padrão.
    """
    # Tenta carregar categorias personalizadas da planilha
    grupos = _carregar_categorias_personalizadas()

    categorias = {}
    nao_classificados = []

    for item in dados:
        nome = item['Despesa'].lower()
        classificado = False
        for grupo, palavras in grupos.items():
            if grupo == 'Outros':
                continue
            for p in palavras:
                if p.lower() in nome:
                    if grupo not in categorias:
                        categorias[grupo] = {mes: 0.0 for mes in MESES_ABREV}
                        categorias[grupo]['Despesa'] = grupo
                    for mes in MESES_ABREV:
                        categorias[grupo][mes] += item[mes]
                    classificado = True
                    break
            if classificado:
                break

        if not classificado:
            nao_classificados.append(item)

    for item in nao_classificados:
        nome_original = item['Despesa']
        if nome_original not in categorias:
            categorias[nome_original] = item.copy()

    return list(categorias.values())


def _carregar_categorias_personalizadas():
    """Tenta ler a planilha 'Categorias' do Excel. Se não existir, usa padrão."""
    padrao = {
        'Moradia': ['aluguel', 'condomínio', 'iptu'],
        'Contas': ['luz', 'água', 'agua', 'internet', 'telefone', 'celular'],
        'Alimentação': ['mercado', 'supermercado', 'feira', 'ifood', 'restaurante', 'alimentação'],
        'Transporte': ['combustível', 'gasolina', 'uber', 'transporte', 'estacionamento', 'ipva'],
        'Saúde': ['plano de saúde', 'farmácia', 'médico', 'dentista', 'academia', 'saúde', 'saude'],
        'Lazer': ['viagem', 'spotify', 'netflix', 'cinema', 'streaming', 'lazer'],
        'Assinaturas': ['spotify', 'google one', 'netflix', 'amazon prime', 'disney', 'hbo'],
        'Educação': ['curso', 'faculdade', 'escola', 'livro'],
        'Vestuário': ['roupa', 'calçado', 'calcado', 'vestuário', 'sapato'],
        'Outros': []
    }

    if not os.path.exists(CAMINHO_EXCEL):
        return padrao

    try:
        wb = openpyxl.load_workbook(CAMINHO_EXCEL, data_only=True)
        if 'Categorias' not in wb.sheetnames:
            wb.close()
            return padrao

        ws = wb['Categorias']
        grupos = {}
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1]:
                grupo = str(row[0]).strip()
                palavras = [str(row[i]).strip().lower() for i in range(1, len(row)) if row[i]]
                if grupo not in grupos:
                    grupos[grupo] = []
                grupos[grupo].extend(palavras)

        wb.close()
        if grupos:
            grupos['Outros'] = []
            return grupos
        return padrao
    except Exception:
        return padrao


@app.route('/')
def home():
    try:
        dados = carregar_dados()
        if not dados:
            return "<h1>⚠️ Nenhum dado encontrado</h1>"

        # Agrupa em categorias para melhor visualização
        categorias = agrupar_categorias(dados)

        # Mês selecionado
        mes_sel = request.args.get('mes', default=obter_mes_atual(), type=str).upper()
        if mes_sel not in MESES_ABREV:
            mes_sel = obter_mes_atual()
        idx_mes = MESES_ABREV.index(mes_sel)
        nome_mes = MESES_NOMES[idx_mes]
        ano_atual = datetime.now().year

        # =============================================
        # TOTAIS MENSAIS (para gráfico de barras e linha)
        # =============================================
        totais_mensais = []
        meses_com_dados = []
        for i, m in enumerate(MESES_ABREV):
            total = sum(item[m] for item in dados)
            if total > 0 or i <= idx_mes:
                totais_mensais.append(round(total, 2))
                meses_com_dados.append(MESES_NOMES[i][:3])

        # =============================================
        # CARDS: Saldo do mês / Despesas do mês
        # =============================================
        total_mes = sum(item[mes_sel] for item in dados)

        # Mês anterior para comparação
        mes_anterior_valor = 0.0
        if idx_mes > 0:
            mes_ant = MESES_ABREV[idx_mes - 1]
            mes_anterior_valor = sum(item[mes_ant] for item in dados)

        variacao = 0.0
        if mes_anterior_valor > 0:
            variacao = ((total_mes - mes_anterior_valor) / mes_anterior_valor) * 100

        # =============================================
        # ANÁLISE MENSAL: barras empilhadas por categoria
        # =============================================
        # Top 5 categorias + Outros para cada mês
        top_cat_nomes = []
        totais_ano_cat = {}
        for cat in categorias:
            nome = cat['Despesa']
            totais_ano_cat[nome] = sum(cat[m] for m in MESES_ABREV)

        top_5 = sorted(totais_ano_cat.items(), key=lambda x: x[1], reverse=True)[:5]
        top_cat_nomes = [t[0] for t in top_5]

        datasets_barras = []
        for i, nome_cat in enumerate(top_cat_nomes):
            valores = []
            for m in MESES_ABREV[:len(meses_com_dados)]:
                for cat in categorias:
                    if cat['Despesa'] == nome_cat:
                        valores.append(round(cat[m], 2))
                        break
                else:
                    valores.append(0)
            datasets_barras.append({
                'label': nome_cat,
                'data': valores,
                'backgroundColor': CORES[i]
            })

        # Outros
        outros_valores = []
        for i_mes, m in enumerate(MESES_ABREV[:len(meses_com_dados)]):
            total_outros = 0.0
            for cat in categorias:
                if cat['Despesa'] not in top_cat_nomes:
                    total_outros += cat[m]
            outros_valores.append(round(total_outros, 2))
        datasets_barras.append({
            'label': 'Outros',
            'data': outros_valores,
            'backgroundColor': '#475569'
        })

        # =============================================
        # DESPESAS NO MÊS: donut chart
        # =============================================
        cats_do_mes = sorted(
            [c for c in categorias if c[mes_sel] > 0],
            key=lambda x: x[mes_sel], reverse=True
        )
        donut_mes_labels = []
        donut_mes_valores = []
        for cat in cats_do_mes[:6]:
            donut_mes_labels.append(cat['Despesa'])
            donut_mes_valores.append(round(cat[mes_sel], 2))
        if len(cats_do_mes) > 6:
            outros = sum(c[mes_sel] for c in cats_do_mes[6:])
            donut_mes_labels.append('Outros')
            donut_mes_valores.append(round(outros, 2))

        pct_mes = (total_mes / sum(totais_mensais) * 100) if sum(totais_mensais) > 0 else 0

        # =============================================
        # DESPESAS NO ANO: donut chart
        # =============================================
        cats_ano = sorted(categorias, key=lambda x: sum(x[m] for m in MESES_ABREV), reverse=True)
        donut_ano_labels = []
        donut_ano_valores = []
        for cat in cats_ano[:6]:
            total_cat = sum(cat[m] for m in MESES_ABREV)
            donut_ano_labels.append(cat['Despesa'])
            donut_ano_valores.append(round(total_cat, 2))
        if len(cats_ano) > 6:
            outros = sum(sum(c[m] for m in MESES_ABREV) for c in cats_ano[6:])
            donut_ano_labels.append('Outros')
            donut_ano_valores.append(round(outros, 2))

        total_ano = sum(donut_ano_valores)

        # =============================================
        # PARTICIPAÇÃO ANUAL: gauge indicators
        # =============================================
        gauges = []
        for cat in cats_ano[:6]:
            total_cat = sum(cat[m] for m in MESES_ABREV)
            pct = (total_cat / total_ano * 100) if total_ano > 0 else 0
            gauges.append({
                'nome': cat['Despesa'],
                'percentual': round(pct, 1)
            })

        # =============================================
        # DETALHAMENTO: horizontal bar (dados brutos)
        # =============================================
        itens_mes = sorted(
            [d for d in dados if d[mes_sel] > 0],
            key=lambda x: x[mes_sel], reverse=True
        )
        detalhamento = []
        for item in itens_mes[:12]:
            detalhamento.append({
                'nome': item['Despesa'],
                'valor': round(item[mes_sel], 2),
                'valor_fmt': formata_real(item[mes_sel])
            })

        # =============================================
        # EVOLUÇÃO MENSAL: line chart (tendência)
        # =============================================
        # Mostrar os últimos meses como linha
        evolucao_valores = totais_mensais[:idx_mes + 1] if len(totais_mensais) > idx_mes else totais_mensais

        # =============================================
        # MONTAR OPÇÕES DE MESES
        # =============================================
        opcoes_meses = []
        for i, abrev in enumerate(MESES_ABREV):
            soma = sum(item[abrev] for item in dados)
            if soma > 0 or i <= idx_mes:
                opcoes_meses.append({
                    "abrev": abrev,
                    "nome": MESES_NOMES[i],
                    "selecionado": (abrev == mes_sel)
                })

        return render_template(
            'index.html',
            # Cards
            mes_nome=nome_mes,
            mes_abrev=mes_sel,
            ano=ano_atual,
            total_mes=formata_real_curto(total_mes),
            total_mes_raw=total_mes,
            total_ano_fmt=formata_real_curto(total_ano),
            total_mes_ant=formata_real_curto(mes_anterior_valor),
            variacao=round(variacao, 1),
            mes_ant_nome=MESES_NOMES[idx_mes - 1] if idx_mes > 0 else '',
            # Sidebar
            opcoes_meses=opcoes_meses,
            meses_abrev=MESES_ABREV,
            meses_nomes=MESES_NOMES,
            # Charts - JSON
            labels_mensais=json.dumps([m[:3] for m in meses_com_dados]),
            totais_mensais=json.dumps(totais_mensais),
            totais_mensais_raw=totais_mensais,
            tamanho_mensais=len(totais_mensais),
            max_mensal=max(totais_mensais) if totais_mensais else 1,
            datasets_barras=json.dumps(datasets_barras),
            donut_mes_labels=json.dumps(donut_mes_labels),
            donut_mes_valores=json.dumps(donut_mes_valores),
            donut_mes_pct=round(pct_mes, 1),
            donut_ano_labels=json.dumps(donut_ano_labels),
            donut_ano_valores=json.dumps(donut_ano_valores),
            donut_ano_pct=round((total_ano / total_ano * 100) if total_ano > 0 else 0, 1),
            gauges=gauges,
            detalhamento=detalhamento,
            evolucao_labels=json.dumps([m[:3] for m in meses_com_dados[:idx_mes + 1]]),
            evolucao_valores=json.dumps(evolucao_valores),
            cores=json.dumps(CORES),
            cores_list=CORES,
            len=len
        )

    except FileNotFoundError as e:
        return f"<h1>⚠️ Arquivo não encontrado</h1><p>{e}</p>", 404
    except Exception as e:
        import traceback
        return (
            f"<h1>❌ Erro interno</h1><p>{e}</p>"
            f"<pre style='background:#1a1a2e;color:#e74c3c;padding:15px;border-radius:8px;"
            f"overflow-x:auto;font-size:12px;max-height:300px;'>{traceback.format_exc()}</pre>",
            500
        )


if __name__ == '__main__':
    app.run(debug=True)